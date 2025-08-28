import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 模板文件路径
template_path = os.path.join(os.path.dirname(__file__), 'sample', '市内交通费明细表-模板.xlsx')

# 加载模板文件
wb = load_workbook(template_path)
ws = wb.active

# 打印工作表名称
print(f"工作表名称: {ws.title}")

# 打印表头
print("表头信息:")
max_row = ws.max_row
max_col = ws.max_column
for col in range(1, max_col + 1):
    col_letter = get_column_letter(col)
    cell_value = ws.cell(row=1, column=col).value
    if cell_value:
        print(f"列 {col_letter}: {cell_value}")

# 打印列宽
print("列宽信息:")
for col in range(1, max_col + 1):
    col_letter = get_column_letter(col)
    width = ws.column_dimensions[col_letter].width
    print(f"列 {col_letter}: {width}")

# 打印一些数据行结构
print("数据行结构示例:")
for row in range(2, min(5, max_row + 1)):
    row_data = []
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        try:
            cell_value = ws.cell(row=row, column=col).value
            row_data.append(str(cell_value) if cell_value else 'None')
        except Exception as e:
            row_data.append(f"Error: {str(e)}")
    print(f"行 {row}: {', '.join(row_data)}")

# 打印合并单元格
print("合并单元格:")
merged_cells = ws.merged_cells.ranges
for merged in merged_cells:
    print(f"{merged}")