import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models.database import get_db, Invoice

class ExcelGenerator:
    """Excel生成器，用于生成符合公司要求的市内交通明细表"""
    def __init__(self):
        # 初始化工作簿和工作表
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "市内交通费明细"

    def _init_table_style(self):
        """初始化表格样式"""
        # 设置标题字体
        title_font = Font(name='宋体', size=14, bold=True)
        # 设置表头字体
        header_font = Font(name='宋体', size=12, bold=True)
        # 设置内容字体
        content_font = Font(name='宋体', size=11)
        # 设置边框
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        # 设置居中对齐
        center_alignment = Alignment(horizontal='center', vertical='center')

        return title_font, header_font, content_font, border, center_alignment

    def generate_transportation_table(self, invoice_ids, additional_info=None):
        """
        生成市内交通明细表
        :param invoice_ids: 选中的发票ID列表
        :param additional_info: 额外信息字典，键为发票ID，值为补充信息
        :return: 生成的Excel文件路径
        """
        # 获取数据库会话和发票数据
        db = next(get_db())
        invoices = db.query(Invoice).filter(Invoice.id.in_(invoice_ids)).all()
        invoices.sort(key=lambda x: x.invoice_date or datetime.date.min)

        # 初始化表格样式
        title_font, header_font, content_font, border, center_alignment = self._init_table_style()

        # 设置标题
        title = "市内交通费明细表"
        self.ws.merge_cells('A1:F1')
        title_cell = self.ws['A1']
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = center_alignment

        # 设置表头
        headers = ["序号", "日期", "出发地", "目的地", "交通方式", "金额", "发票编号", "备注"]
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            cell = self.ws[f'{col_letter}2']
            cell.value = header
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            # 设置列宽
            self.ws.column_dimensions[col_letter].width = 15

        # 填充数据
        for row_idx, invoice in enumerate(invoices, 3):
            # 序号
            self.ws[f'A{row_idx}'].value = row_idx - 2
            # 日期
            if invoice.invoice_date:
                self.ws[f'B{row_idx}'].value = invoice.invoice_date.strftime('%Y-%m-%d')
            # 出发地和目的地（需要用户补充，这里用默认值）
            self.ws[f'C{row_idx}'].value = "公司"
            self.ws[f'D{row_idx}'].value = "办事地点"
            # 交通方式（根据发票类型自动判断或用户补充）
            self.ws[f'E{row_idx}'].value = "出租车" if "滴滴" in (invoice.recognized_text or "") else "其他"
            # 金额
            self.ws[f'F{row_idx}'].value = invoice.amount
            # 发票编号
            self.ws[f'G{row_idx}'].value = invoice.invoice_number
            # 备注
            if additional_info and invoice.id in additional_info:
                self.ws[f'H{row_idx}'].value = additional_info[invoice.id]
            else:
                self.ws[f'H{row_idx}'].value = ""

            # 设置单元格样式
            for col_idx in range(1, 9):
                col_letter = get_column_letter(col_idx)
                cell = self.ws[f'{col_letter}{row_idx}']
                cell.font = content_font
                cell.border = border
                cell.alignment = center_alignment

        # 设置金额合计
        total_row = len(invoices) + 3
        self.ws.merge_cells(f'A{total_row}:E{total_row}')
        self.ws[f'A{total_row}'].value = "合计"
        self.ws[f'A{total_row}'].font = header_font
        self.ws[f'A{total_row}'].alignment = center_alignment
        self.ws[f'A{total_row}'].border = border

        # 计算总金额
        total_amount = sum(invoice.amount for invoice in invoices if invoice.amount)
        self.ws[f'F{total_row}'].value = total_amount
        self.ws[f'F{total_row}'].font = header_font
        self.ws[f'F{total_row}'].alignment = center_alignment
        self.ws[f'F{total_row}'].border = border

        # 保存文件
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output')
        os.makedirs(output_dir, exist_ok=True)
        filename = f"市内交通费明细表_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(output_dir, filename)
        self.wb.save(output_path)

        return output_path

    def generate_from_category(self, category):
        """
        根据分类生成市内交通明细表
        :param category: 发票分类
        :return: 生成的Excel文件路径
        """
        # 获取数据库会话和发票数据
        db = next(get_db())
        invoices = db.query(Invoice).filter(Invoice.category == category).all()
        invoice_ids = [invoice.id for invoice in invoices]

        return self.generate_transportation_table(invoice_ids)